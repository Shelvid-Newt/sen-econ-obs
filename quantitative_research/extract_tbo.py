#!/usr/bin/env python3
"""
ETL Pipeline for Senegal Economic Observatory (SEO)
Extracts macroeconomic time series from the administrative TBO Excel workbook.
Author: DataPipelineArchitect
Date: May 2026
"""

import os
import json
import re
from datetime import datetime
import openpyxl

# We import stats functions from pure Python to keep it light and bulletproof
# but if numpy is available we can use it. We'll implement a clean, pure Python fallback.

def median(lst):
    """Compute median of a list of numbers."""
    if not lst:
        return 0.0
    sorted_lst = sorted(lst)
    n = len(sorted_lst)
    if n % 2 == 1:
        return float(sorted_lst[n // 2])
    else:
        return float((sorted_lst[n // 2 - 1] + sorted_lst[n // 2]) / 2.0)

def detect_outliers_mad(values, threshold=3.5):
    """
    Detect outliers using Median Absolute Deviation (MAD).
    Returns a list of booleans indicating whether each value is an outlier.
    """
    # Filter out non-numeric values
    clean_vals = [v for v in values if v is not None and not isinstance(v, str)]
    if len(clean_vals) < 3:
        return [False] * len(values)
        
    med = median(clean_vals)
    abs_deviations = [abs(v - med) for v in clean_vals]
    mad = median(abs_deviations)
    
    if mad == 0.0:
        # Fallback to mean absolute deviation if MAD is 0 (all values are identical)
        mean_val = sum(clean_vals) / len(clean_vals)
        mean_abs_dev = sum(abs(v - mean_val) for v in clean_vals) / len(clean_vals)
        mad = mean_abs_dev
        if mad == 0.0:
            return [False] * len(values)
            
    outliers = []
    for v in values:
        if v is None or isinstance(v, str):
            outliers.append(False)
            continue
        dev = abs(v - med)
        # Modified Z-score approximation using 0.6745
        modified_z = 0.6745 * dev / mad
        outliers.append(modified_z > threshold)
        
    return outliers

def interpolate_series(values, outliers_flags):
    """
    Interpolate missing values and outliers in a series.
    Returns:
      - cleaned_values: list of floats
      - imputed_flags: list of booleans
    """
    n = len(values)
    cleaned = []
    
    # 1. First pass: Convert strings, None, and outliers to None
    for i, v in enumerate(values):
        if v is None or isinstance(v, str) or outliers_flags[i]:
            cleaned.append(None)
        else:
            try:
                cleaned.append(float(v))
            except (ValueError, TypeError):
                cleaned.append(None)
                
    # Find indices of valid values
    valid_idx = [i for i, v in enumerate(cleaned) if v is not None]
    
    if not valid_idx:
        # If the entire series is empty, return 0.0
        return [0.0] * n, [True] * n
        
    imputed_flags = [False] * n
    result = list(cleaned)
    
    # 2. Second pass: Perform linear interpolation
    for i in range(n):
        if result[i] is None:
            imputed_flags[i] = True
            left = [idx for idx in valid_idx if idx < i]
            right = [idx for idx in valid_idx if idx > i]
            
            if left and right:
                # Normal linear interpolation between nearest neighbors
                l_idx = left[-1]
                r_idx = right[0]
                l_val = cleaned[l_idx]
                r_val = cleaned[r_idx]
                result[i] = l_val + (r_val - l_val) * (i - l_idx) / (r_idx - l_idx)
            elif left:
                # Extrapolate using last valid value (forward fill)
                result[i] = cleaned[left[-1]]
            elif right:
                # Extrapolate using first valid value (backward fill)
                result[i] = cleaned[right[0]]
                
    return result, imputed_flags

class TBOPipeline:
    def __init__(self, file_path):
        self.file_path = file_path
        print(f"Initializing SEO ETL Pipeline. Loading workbook: {file_path}")
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        
    def get_sheet_series(self, sheet_name, date_col, val_col, start_row, end_row):
        """Helper to extract raw values and dates from a sheet column."""
        sheet = self.wb[sheet_name]
        dates = []
        raw_vals = []
        for r in range(start_row, end_row + 1):
            dt = sheet.cell(row=r, column=date_col).value
            val = sheet.cell(row=r, column=val_col).value
            
            # Format date
            date_str = None
            if dt is not None:
                if isinstance(dt, datetime) or hasattr(dt, 'year'):
                    date_str = dt.strftime("%Y-%m-%d")
                else:
                    # try parsing
                    dt_clean = str(dt).strip().split(" ")[0]
                    if re.match(r"\d{4}-\d{2}-\d{2}", dt_clean):
                        date_str = dt_clean
            
            if date_str:
                dates.append(date_str)
                # Clean value if string dash
                if val is not None and str(val).strip() == "-":
                    raw_vals.append(None)
                else:
                    raw_vals.append(val)
                    
        return dates, raw_vals

    def process_series(self, sheet_name, date_col, val_col, start_row, end_row, threshold=3.5, filter_zeros_pre=False):
        """Extract raw values from a sheet column. Only interpolate genuine missing values (None, '-').
        Outlier detection has been disabled: DPEE official data should be taken as-is."""
        dates, raw_vals = self.get_sheet_series(sheet_name, date_col, val_col, start_row, end_row)
        
        # For gold production or columns where leading zeros are expected before mine starts
        if filter_zeros_pre:
            first_non_null_idx = None
            for idx, val in enumerate(raw_vals):
                if val is not None and not isinstance(val, str) and float(val) > 0:
                    first_non_null_idx = idx
                    break
            if first_non_null_idx is not None:
                for idx in range(first_non_null_idx):
                    if raw_vals[idx] is None or str(raw_vals[idx]).strip() == "-":
                        raw_vals[idx] = 0.0

        # NO outlier detection — official DPEE data is authoritative
        no_outliers = [False] * len(raw_vals)
        
        # Interpolate only genuine missing values (None, non-numeric)
        imputed_vals, imputed_flags = interpolate_series(raw_vals, no_outliers)
        
        series_data = {}
        for i in range(len(dates)):
            series_data[dates[i]] = {
                "value": imputed_vals[i],
                "raw_value": raw_vals[i],
                "imputed": imputed_flags[i],
                "outlier": False
            }
        return series_data

    def run(self, output_dir):
        os.makedirs(output_dir, exist_ok=True)
        print(f"Data will be exported to: {output_dir}")
        
        consolidated = {}
        
        # 1. Extract Sectoral Indices (Primaire, Secondaire, Tertiaire, Commerce)
        # The "graphes" sheet aggregates formulas referencing source sheets, but is
        # fragile: when a new month is appended, the formula author has occasionally
        # shifted by one row (e.g. graphes!E122 pointed at 'Indice Secondaire'!B128
        # for Jan 2026 instead of B127, yielding 0). We bypass `graphes` and read
        # each sector directly from its authoritative source sheet.
        # - primaire_brute  ← Primaire!H              (rows 7-247, date in col A)
        # - secondaire_brute ← Indice Secondaire!B    (rows 7-127, date in col A)
        # - tertiaire_brute ← ICAS et ICAC!C          (rows 7-247, date in col A)
        # - commerce_brute  ← ICAS et ICAC!AO         (rows 7-247, date in col A)
        # CVS variants remain broken (#REF! in source) — emit explicit nulls.
        print("Extracting Sectoral Indices...")
        # All four indices are anchored to Jan 2016 (base 100 = 2022). The source
        # sheets store much older history (back to 2006/2007) in earlier rows, but
        # the H/C/AO index columns are empty before 2016 — extracting them would
        # poison the interpolation. We restrict each range to its Jan-2016 anchor.
        sectors_sources = {
            "primaire_brute":   ("Primaire",          8, 127, 247),  # col H,  rows 127-247
            "secondaire_brute": ("Indice Secondaire", 2,   7, 127),  # col B,  rows   7-127
            "tertiaire_brute":  ("ICAS et ICAC",      3, 115, 235),  # col C,  rows 115-235
            "commerce_brute":   ("ICAS et ICAC",     41, 115, 235),  # col AO, rows 115-235
        }
        cvs_names = ["primaire_cvs", "secondaire_cvs", "tertiaire_cvs", "commerce_cvs"]
        sectors_data = {}
        for name, (sheet, col, sr, er) in sectors_sources.items():
            sectors_data[name] = self.process_series(sheet, 1, col, sr, er)
        # Build empty CVS series mirroring the dates of their _brute counterpart
        for cvs_name in cvs_names:
            brute_name = cvs_name.replace("_cvs", "_brute")
            sectors_data[cvs_name] = {
                date: {
                    "value": None,
                    "raw_value": None,
                    "imputed": False,
                    "outlier": False,
                    "status": "broken_formula_in_source"
                }
                for date in sectors_data[brute_name].keys()
            }
        # Preserve the original key order expected by downstream consumers
        sectors_mapping = {
            "primaire_brute": None, "primaire_cvs": None,
            "secondaire_brute": None, "secondaire_cvs": None,
            "tertiaire_brute": None, "tertiaire_cvs": None,
            "commerce_brute": None, "commerce_cvs": None,
        }
        sectors_data = {k: sectors_data[k] for k in sectors_mapping.keys()}
                
        # 2. Extract Or Production
        # sheet: Secteur Secondaire, date: row 7-247, Column 10
        print("Extracting Or Production...")
        or_data = self.process_series("Secteur Secondaire", 1, 10, 7, 247, filter_zeros_pre=True)
        
        # 3. Extract Ciment Series (Production, Ventes Locales, Exportations)
        # sheet: Secteur Secondaire, date: row 7-247. Columns are R/S/T (18/19/20):
        # col 17 is an off-topic index (~19) — using it produced a bogus "kt"
        # headline. The real volumes in kt are Production=col 18, Ventes locales=
        # col 19, Exportations=col 20 (verified against the Jan-2026 TBO).
        print("Extracting Ciment Series...")
        ciment_data = {
            "production": self.process_series("Secteur Secondaire", 1, 18, 7, 247),
            "ventes_locales": self.process_series("Secteur Secondaire", 1, 19, 7, 247),
            "exportations": self.process_series("Secteur Secondaire", 1, 20, 7, 247)
        }
        
        # 4. Extract Brent Oil Price
        # sheet: Environnement International, date: row 7-247, Column 7
        print("Extracting Brent oil price...")
        brent_data = self.process_series("Environnement International", 1, 7, 7, 247)

        # 4b. Extract Exchange Rates (same sheet): EUR/USD col 11, USD/CFA col 10
        print("Extracting Exchange Rates (EUR/USD, USD/CFA)...")
        change_data = {
            "eur_usd": self.process_series("Environnement International", 1, 11, 7, 247),
            "usd_cfa": self.process_series("Environnement International", 1, 10, 7, 247),
        }

        # 5. Extract Recettes Fiscales
        # sheet: Finances Publiques , date: row 7-247, Column 4
        print("Extracting Recettes Fiscales...")
        recettes_fiscales_data = self.process_series("Finances Publiques ", 1, 4, 7, 247)

        # 5b. Finances detail: TVA(7), Douanes(8), Recettes non fiscales(11),
        # Masse salariale(12), Effectifs(14). (No expenditure/debt block in TBO.)
        print("Extracting Finances detail...")
        finances_data = {
            "tva": self.process_series("Finances Publiques ", 1, 7, 7, 247),
            "douanes": self.process_series("Finances Publiques ", 1, 8, 7, 247),
            "recettes_non_fiscales": self.process_series("Finances Publiques ", 1, 11, 7, 247),
            "masse_salariale": self.process_series("Finances Publiques ", 1, 12, 7, 247),
            "effectifs": self.process_series("Finances Publiques ", 1, 14, 7, 247),
        }
        
        # 6. Extract Trafic Maritime
        # sheet: Transport, date: row 7-247
        print("Extracting Maritime Traffic...")
        trafic_maritime_data = {
            "embarquements_total": self.process_series("Transport", 1, 10, 7, 247),
            "debarquements_total": self.process_series("Transport", 1, 14, 7, 247)
        }

        # 6b. Extract Air Traffic (Transport sheet): mouvements(2), passagers(3), fret(7)
        print("Extracting Air Traffic...")
        transport_aerien_data = {
            "mouvements": self.process_series("Transport", 1, 2, 7, 247),
            "passagers": self.process_series("Transport", 1, 3, 7, 247),
            "fret": self.process_series("Transport", 1, 7, 7, 247),
        }
        
        # 7. Extract Pêche Series (Landings from Primaire and Exports from Commerce Extérieur)
        print("Extracting Pêche data...")
        peche_data = {
            "debarquements_total": self.process_series("Primaire", 1, 3, 7, 247),
            "debarquements_industrielle": self.process_series("Primaire", 1, 4, 7, 247),
            "debarquements_artisanale": self.process_series("Primaire", 1, 5, 7, 247),
            "exportations_volume": self.process_series("Commerce Extérieur", 1, 8, 7, 247)
        }

        # 7b. Extract Foreign Trade flows (Commerce Extérieur), 1000 tonnes
        #  Imports: Riz(2), Blé(3), Pétrole Brut(4), Hydrocarbures raffinés(5)
        #  Exports: Arachide vol(6) / valeur Mds FCFA(7), Pêche(8)
        print("Extracting Commerce Extérieur flows (échanges)...")
        echanges_data = {
            "imports": {
                "riz": self.process_series("Commerce Extérieur", 1, 2, 7, 247),
                "ble": self.process_series("Commerce Extérieur", 1, 3, 7, 247),
                "petrole_brut": self.process_series("Commerce Extérieur", 1, 4, 7, 247),
                "hydrocarbures_raffines": self.process_series("Commerce Extérieur", 1, 5, 7, 247),
            },
            "exports": {
                "arachide": self.process_series("Commerce Extérieur", 1, 6, 7, 247),
                "peche": self.process_series("Commerce Extérieur", 1, 8, 7, 247),
            },
            "arachide_valeur": self.process_series("Commerce Extérieur", 1, 7, 7, 247),
        }

        # 8. Extract Prix Vivriers Series (National Averages)
        # sheet: Prix Intérieurs, date: row 7-211
        print("Extracting Prix Vivriers Series...")
        prix_vivriers_mapping = {
            "oignon_local": 2,
            "oignon_importe": 3,
            "millet_souna": 4,
            "sorgho": 5,
            "mais": 6,
            "riz_brise_parfume_luxe": 7,
            "riz_brise_parfume_ordinaire": 8,
            "riz_brise_non_parfume": 9,
            "riz_brise_indien_ordinaire": 10,
            "riz_brise_local": 11
        }
        prix_vivriers_data = {}
        for name, col in prix_vivriers_mapping.items():
            prix_vivriers_data[name] = self.process_series("Prix Intérieurs", 1, col, 7, 211, threshold=3.0)

        # 9. Extract Regional Price Blocks
        print("Extracting Regional Price snapshots...")
        sheet_prix = self.wb["Prix Intérieurs"]
        regions_order = [
            "Dakar", "Thiès", "Diourbel", "Louga", "St louis", 
            "Matam", "Fatick", "Kaolack", "Tamba", "Kolda", "Ziguinchor", "Prix Moyen"
        ]
        products_order = [
            "oignon_local", "oignon_importe", "millet_souna", "sorgho", "mais",
            "riz_brise_parfume_luxe", "riz_brise_parfume_ordinaire", 
            "riz_brise_non_parfume", "riz_brise_indien_ordinaire", "riz_brise_local"
        ]
        
        regional_price_blocks = {}
        
        # Scan column 14 (N) for year markers and block boundaries
        for r in range(1, sheet_prix.max_row + 1):
            val = sheet_prix.cell(row=r, column=14).value
            val_str = str(val).strip() if val is not None else ""
            
            is_block_start = False
            block_label = None
            value_start_row = None
            
            if val_str in ["2009", "2010", "2011", "2012", "2013", "2014", "2015", "2016", "2017", "2018", "2019"]:
                is_block_start = True
                block_label = f"Janvier {val_str}"
                value_start_row = r + 7
                
                # Adjustments for Février and Mars 2019
                if r == 242:
                    block_label = "Février 2019"
                    value_start_row = 249
                elif r == 262:
                    block_label = "Mars 2019"
                    value_start_row = 268
                    
            if is_block_start and value_start_row <= sheet_prix.max_row:
                block_data = {}
                for i, region in enumerate(regions_order):
                    row_idx = value_start_row + i
                    region_prices = {}
                    for col_offset, prod in enumerate(products_order):
                        col_idx = 15 + col_offset  # Column O is 15
                        cell_val = sheet_prix.cell(row=row_idx, column=col_idx).value
                        if cell_val is None or str(cell_val).strip() in ["-", ""]:
                            region_prices[prod] = None
                        else:
                            try:
                                region_prices[prod] = float(cell_val)
                            except ValueError:
                                region_prices[prod] = None
                    block_data[region] = region_prices
                regional_price_blocks[block_label] = block_data
                
        # Consolidated Time Series Builder
        # Gather all unique dates across series
        all_dates = set()
        all_dates.update(sectors_data["primaire_brute"].keys())
        all_dates.update(or_data.keys())
        all_dates.update(ciment_data["production"].keys())
        all_dates.update(brent_data.keys())
        all_dates.update(recettes_fiscales_data.keys())
        all_dates.update(trafic_maritime_data["embarquements_total"].keys())
        all_dates.update(peche_data["debarquements_total"].keys())
        all_dates.update(prix_vivriers_data["oignon_local"].keys())
        
        sorted_dates = sorted(list(all_dates))
        print(f"Building consolidated time series. Total months covered: {len(sorted_dates)} (from {sorted_dates[0]} to {sorted_dates[-1]})")
        
        consolidated_series = []
        for date_str in sorted_dates:
            dt_record = {"date": date_str}
            
            # Sectors
            dt_record["secteurs"] = {}
            for name in sectors_mapping.keys():
                if date_str in sectors_data[name]:
                    dt_record["secteurs"][name] = sectors_data[name][date_str]
                else:
                    dt_record["secteurs"][name] = None
                    
            # Or
            if date_str in or_data:
                dt_record["or"] = or_data[date_str]
            else:
                dt_record["or"] = None
                
            # Ciment
            dt_record["ciment"] = {}
            for name in ["production", "ventes_locales", "exportations"]:
                if date_str in ciment_data[name]:
                    dt_record["ciment"][name] = ciment_data[name][date_str]
                else:
                    dt_record["ciment"][name] = None
                    
            # Brent
            if date_str in brent_data:
                dt_record["brent"] = brent_data[date_str]
            else:
                dt_record["brent"] = None
                
            # Recettes Fiscales
            if date_str in recettes_fiscales_data:
                dt_record["recettes_fiscales"] = recettes_fiscales_data[date_str]
            else:
                dt_record["recettes_fiscales"] = None
                
            # Trafic Maritime
            dt_record["trafic_maritime"] = {}
            for name in ["embarquements_total", "debarquements_total"]:
                if date_str in trafic_maritime_data[name]:
                    dt_record["trafic_maritime"][name] = trafic_maritime_data[name][date_str]
                else:
                    dt_record["trafic_maritime"][name] = None
                    
            # Pêche
            dt_record["peche"] = {}
            for name in ["debarquements_total", "debarquements_industrielle", "debarquements_artisanale", "exportations_volume"]:
                if date_str in peche_data[name]:
                    dt_record["peche"][name] = peche_data[name][date_str]
                else:
                    dt_record["peche"][name] = None
                    
            # Prix Vivriers
            dt_record["prix_vivriers"] = {}
            for name in prix_vivriers_mapping.keys():
                if date_str in prix_vivriers_data[name]:
                    dt_record["prix_vivriers"][name] = prix_vivriers_data[name][date_str]
                else:
                    dt_record["prix_vivriers"][name] = None
                    
            consolidated_series.append(dt_record)

        # WRITE JSON FILES
        def save_json(data, filename):
            file_path = os.path.join(output_dir, filename)
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"Successfully exported: {file_path}")

        save_json(sectors_data, "sectors.json")
        save_json(or_data, "or.json")
        save_json(ciment_data, "ciment.json")
        save_json(brent_data, "brent.json")
        save_json(change_data, "change.json")
        save_json(recettes_fiscales_data, "recettes_fiscales.json")
        save_json(finances_data, "finances.json")
        save_json(trafic_maritime_data, "trafic_maritime.json")
        save_json(transport_aerien_data, "transport_aerien.json")
        save_json(peche_data, "peche.json")
        save_json(echanges_data, "echanges.json")
        save_json(prix_vivriers_data, "prix_vivriers.json")
        save_json(regional_price_blocks, "prix_regionaux.json")
        save_json(consolidated_series, "consolidated.json")
        
        print("\nAll data ingestion, cleanup, outlier detection, and interpolation completed successfully!")

if __name__ == "__main__":
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    wb_path = os.path.join(script_dir, "..", "raw_data", "TBO_JANVIER_2026.xlsx")
    out_dir = os.path.join(script_dir, "..", "frontend", "public", "data")
    
    pipeline = TBOPipeline(wb_path)
    pipeline.run(out_dir)
