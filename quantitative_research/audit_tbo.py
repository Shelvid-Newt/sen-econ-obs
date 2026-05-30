#!/usr/bin/env python3
"""
AUDIT SCRIPT — Cross-validate extracted JSON against the raw Excel TBO.
Prints a report of discrepancies for the most recent month (January 2026).

Usage: python audit_tbo.py
"""

import json
import os
import openpyxl

WB_PATH = r"c:\Users\david\OneDrive\Desktop\senegal-economic-observatory\bulk data\TBO_JANVIER_2026 (1).xlsx"
DATA_DIR = r"c:\Users\david\OneDrive\Desktop\senegal-economic-observatory\frontend\public\data"

def load_json(name):
    with open(os.path.join(DATA_DIR, name), "r", encoding="utf-8") as f:
        return json.load(f)

def cell(ws, row, col):
    """Get a cell value, return None for blanks and dashes."""
    v = ws.cell(row=row, column=col).value
    if v is None or str(v).strip() in ["-", ""]:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return v

def fmt(v):
    if v is None:
        return "None"
    if isinstance(v, float):
        return f"{v:.4f}"
    return str(v)

def compare(label, json_val, excel_val, tolerance=0.5):
    """Compare two values, return (ok, message)."""
    if json_val is None and excel_val is None:
        return True, f"  ✅ {label}: Both None"
    if json_val is None and excel_val is not None:
        return False, f"  ❌ {label}: JSON=None, Excel={fmt(excel_val)}"
    if json_val is not None and excel_val is None:
        return False, f"  ⚠️  {label}: JSON={fmt(json_val)}, Excel=None (missing in source?)"
    try:
        jv = float(json_val)
        ev = float(excel_val)
        diff = abs(jv - ev)
        pct = (diff / abs(ev)) * 100 if ev != 0 else (0 if diff == 0 else 100)
        if pct <= tolerance:
            return True, f"  ✅ {label}: JSON={fmt(jv)}, Excel={fmt(ev)} (diff={pct:.2f}%)"
        else:
            return False, f"  ❌ {label}: JSON={fmt(jv)}, Excel={fmt(ev)} — DIFF={pct:.1f}%"
    except (ValueError, TypeError):
        eq = str(json_val) == str(excel_val)
        return eq, f"  {'✅' if eq else '❌'} {label}: JSON={json_val}, Excel={excel_val}"

def main():
    print("=" * 70)
    print("AUDIT — Senegal Economic Observatory Data Pipeline")
    print("Cross-validating JSON outputs against raw TBO Excel")
    print("=" * 70)
    
    wb = openpyxl.load_workbook(WB_PATH, data_only=True)
    
    # List all sheets
    print(f"\nSheets in workbook: {wb.sheetnames}\n")
    
    errors = 0
    checks = 0
    
    # ── FOCUS MONTH: January 2026 = 2026-01-01 ──
    TARGET = "2026-01-01"
    
    # ─────────────────────────────────────────────
    # 1. SECTORS — Indices d'activité
    # ─────────────────────────────────────────────
    print("\n─── SECTEURS D'ACTIVITÉ ───")
    sectors = load_json("sectors.json")
    
    # Primaire: sheet "Primaire", col H=8, need to find the row for Jan 2026
    ws_prim = wb["Primaire"]
    for r in range(7, 250):
        d = ws_prim.cell(row=r, column=1).value
        if d and hasattr(d, 'year') and d.year == 2026 and d.month == 1:
            excel_prim = cell(ws_prim, r, 8)
            json_prim = sectors.get("primaire_brute", {}).get(TARGET, {})
            jv = json_prim.get("value") if json_prim else None
            ok, msg = compare("Primaire brute (Jan 2026)", jv, excel_prim)
            print(msg)
            checks += 1; errors += 0 if ok else 1
            break
    
    # Secondaire: sheet "Indice Secondaire", col B=2
    ws_sec = wb["Indice Secondaire"]
    for r in range(7, 130):
        d = ws_sec.cell(row=r, column=1).value
        if d and hasattr(d, 'year') and d.year == 2026 and d.month == 1:
            excel_sec = cell(ws_sec, r, 2)
            json_sec = sectors.get("secondaire_brute", {}).get(TARGET, {})
            jv = json_sec.get("value") if json_sec else None
            ok, msg = compare("Secondaire brute (Jan 2026)", jv, excel_sec)
            print(msg)
            checks += 1; errors += 0 if ok else 1
            break
    
    # Tertiaire: sheet "ICAS et ICAC", col C=3
    ws_icas = wb["ICAS et ICAC"]
    for r in range(7, 240):
        d = ws_icas.cell(row=r, column=1).value
        if d and hasattr(d, 'year') and d.year == 2026 and d.month == 1:
            excel_tert = cell(ws_icas, r, 3)
            json_tert = sectors.get("tertiaire_brute", {}).get(TARGET, {})
            jv = json_tert.get("value") if json_tert else None
            ok, msg = compare("Tertiaire brute (Jan 2026)", jv, excel_tert)
            print(msg)
            checks += 1; errors += 0 if ok else 1
            
            # Commerce: same sheet, col AO=41
            excel_com = cell(ws_icas, r, 41)
            json_com = sectors.get("commerce_brute", {}).get(TARGET, {})
            jv = json_com.get("value") if json_com else None
            ok, msg = compare("Commerce brute (Jan 2026)", jv, excel_com)
            print(msg)
            checks += 1; errors += 0 if ok else 1
            break
    
    # ─────────────────────────────────────────────
    # 2. OR — Production aurifère
    # ─────────────────────────────────────────────
    print("\n─── OR INDUSTRIEL ───")
    or_data = load_json("or.json")
    ws_ss = wb["Secteur Secondaire"]
    for r in range(7, 250):
        d = ws_ss.cell(row=r, column=1).value
        if d and hasattr(d, 'year') and d.year == 2026 and d.month == 1:
            excel_or = cell(ws_ss, r, 10)
            json_or = or_data.get(TARGET, {})
            jv = json_or.get("value") if json_or else None
            ok, msg = compare("Or production kg (Jan 2026)", jv, excel_or)
            print(msg)
            checks += 1; errors += 0 if ok else 1
            break
    
    # ─────────────────────────────────────────────
    # 3. CIMENT
    # ─────────────────────────────────────────────
    print("\n─── CIMENT ───")
    ciment = load_json("ciment.json")
    for r in range(7, 250):
        d = ws_ss.cell(row=r, column=1).value
        if d and hasattr(d, 'year') and d.year == 2026 and d.month == 1:
            for col, name in [(18, "production"), (19, "ventes_locales"), (20, "exportations")]:
                excel_v = cell(ws_ss, r, col)
                json_v = ciment.get(name, {}).get(TARGET, {})
                jv = json_v.get("value") if json_v else None
                ok, msg = compare(f"Ciment {name} (Jan 2026)", jv, excel_v)
                print(msg)
                checks += 1; errors += 0 if ok else 1
            break
    
    # ─────────────────────────────────────────────
    # 4. BRENT + CHANGE
    # ─────────────────────────────────────────────
    print("\n─── ENVIRONNEMENT INTERNATIONAL ───")
    brent = load_json("brent.json")
    change = load_json("change.json")
    ws_ei = wb["Environnement International"]
    for r in range(7, 250):
        d = ws_ei.cell(row=r, column=1).value
        if d and hasattr(d, 'year') and d.year == 2026 and d.month == 1:
            # Brent col 7
            excel_brent = cell(ws_ei, r, 7)
            json_brent = brent.get(TARGET, {})
            jv = json_brent.get("value") if json_brent else None
            ok, msg = compare("Brent $/baril (Jan 2026)", jv, excel_brent)
            print(msg)
            checks += 1; errors += 0 if ok else 1
            
            # USD/CFA col 10
            excel_usdcfa = cell(ws_ei, r, 10)
            json_usdcfa = change.get("usd_cfa", {}).get(TARGET, {})
            jv = json_usdcfa.get("value") if json_usdcfa else None
            ok, msg = compare("USD/CFA (Jan 2026)", jv, excel_usdcfa)
            print(msg)
            checks += 1; errors += 0 if ok else 1
            
            # EUR/USD col 11
            excel_eurusd = cell(ws_ei, r, 11)
            json_eurusd = change.get("eur_usd", {}).get(TARGET, {})
            jv = json_eurusd.get("value") if json_eurusd else None
            ok, msg = compare("EUR/USD (Jan 2026)", jv, excel_eurusd)
            print(msg)
            checks += 1; errors += 0 if ok else 1
            break
    
    # ─────────────────────────────────────────────
    # 5. RECETTES FISCALES
    # ─────────────────────────────────────────────
    print("\n─── FINANCES PUBLIQUES ───")
    recettes = load_json("recettes_fiscales.json")
    ws_fp = wb["Finances Publiques "]
    for r in range(7, 250):
        d = ws_fp.cell(row=r, column=1).value
        if d and hasattr(d, 'year') and d.year == 2026 and d.month == 1:
            excel_rec = cell(ws_fp, r, 4)
            json_rec = recettes.get(TARGET, {})
            jv = json_rec.get("value") if json_rec else None
            ok, msg = compare("Recettes fiscales Mds (Jan 2026)", jv, excel_rec)
            print(msg)
            checks += 1; errors += 0 if ok else 1
            break
    
    # ─────────────────────────────────────────────
    # 6. TRANSPORT MARITIME
    # ─────────────────────────────────────────────
    print("\n─── TRANSPORT MARITIME ───")
    maritime = load_json("trafic_maritime.json")
    ws_tr = wb["Transport"]
    for r in range(7, 250):
        d = ws_tr.cell(row=r, column=1).value
        if d and hasattr(d, 'year') and d.year == 2026 and d.month == 1:
            excel_emb = cell(ws_tr, r, 10)
            json_emb = maritime.get("embarquements_total", {}).get(TARGET, {})
            jv = json_emb.get("value") if json_emb else None
            ok, msg = compare("Embarquements total (Jan 2026)", jv, excel_emb)
            print(msg)
            checks += 1; errors += 0 if ok else 1
            
            excel_deb = cell(ws_tr, r, 14)
            json_deb = maritime.get("debarquements_total", {}).get(TARGET, {})
            jv = json_deb.get("value") if json_deb else None
            ok, msg = compare("Débarquements total (Jan 2026)", jv, excel_deb)
            print(msg)
            checks += 1; errors += 0 if ok else 1
            break
    
    # ─────────────────────────────────────────────
    # 7. PÊCHE
    # ─────────────────────────────────────────────
    print("\n─── PÊCHE ───")
    peche = load_json("peche.json")
    for r in range(7, 250):
        d = ws_prim.cell(row=r, column=1).value
        if d and hasattr(d, 'year') and d.year == 2026 and d.month == 1:
            for col, name in [(3, "debarquements_total"), (4, "debarquements_industrielle"), (5, "debarquements_artisanale")]:
                excel_v = cell(ws_prim, r, col)
                json_v = peche.get(name, {}).get(TARGET, {})
                jv = json_v.get("value") if json_v else None
                ok, msg = compare(f"Pêche {name} (Jan 2026)", jv, excel_v)
                print(msg)
                checks += 1; errors += 0 if ok else 1
            break
    
    # ─────────────────────────────────────────────
    # 8. SPOT-CHECK: Dec 2025 and a few older months
    # ─────────────────────────────────────────────
    print("\n─── SPOT-CHECKS (Dec 2025, Jul 2025) ───")
    spot_months = [("2025-12-01", 2025, 12), ("2025-07-01", 2025, 7)]
    for target, yr, mo in spot_months:
        # Check Brent
        for r in range(7, 250):
            d = ws_ei.cell(row=r, column=1).value
            if d and hasattr(d, 'year') and d.year == yr and d.month == mo:
                excel_brent = cell(ws_ei, r, 7)
                json_brent = brent.get(target, {})
                jv = json_brent.get("value") if json_brent else None
                ok, msg = compare(f"Brent ({target[:7]})", jv, excel_brent)
                print(msg)
                checks += 1; errors += 0 if ok else 1
                break
        
        # Check Or
        for r in range(7, 250):
            d = ws_ss.cell(row=r, column=1).value
            if d and hasattr(d, 'year') and d.year == yr and d.month == mo:
                excel_or = cell(ws_ss, r, 10)
                json_or = or_data.get(target, {})
                jv = json_or.get("value") if json_or else None
                ok, msg = compare(f"Or ({target[:7]})", jv, excel_or)
                print(msg)
                checks += 1; errors += 0 if ok else 1
                break
    
    # ─────────────────────────────────────────────
    # SUMMARY
    # ─────────────────────────────────────────────
    print("\n" + "=" * 70)
    print(f"AUDIT COMPLETE: {checks} checks, {errors} discrepancies found")
    if errors == 0:
        print("✅ All values match the Excel source within tolerance (0.5%)")
    else:
        print(f"⚠️  {errors} value(s) need investigation")
    print("=" * 70)

if __name__ == "__main__":
    main()
